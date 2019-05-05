from MinGen.dataclasses import PromRow
import openpyxl


class MinGen:
    def __init__(self, signs_count, objects_count, data):
        self.signs_count = 2 * signs_count
        self.objects_count = objects_count
        self.data = self._form_data(data)
        self.prom_tab = self._make_prom_tab()
        self.it = '0' * self.signs_count
        self.result = []

    @staticmethod
    def _form_data(data):
        formed_data = []
        for i in data:
            formed_row = []
            for j in i:
                formed_row.append(j)
                formed_row.append(0 if j else 1)
            formed_data.append(formed_row)
        return formed_data

    def print_data(self):
        print(' '.join(self.chars))
        for i in range(self.objects_count):
            print(' '.join([str(j) for j in self.data[i]]))

    def print_result(self):
        for res in self.result:
            print(res)

    def print_prom_tab(self):
        data_to_print = [['X', 'K', 'X\'', 'X"']]
        for i in self.prom_tab:
            data_to_print.append([i.X_name, '+' if i.Key else '-', ''.join([str(j) for j in i.X_1]), i.X_2])
        for row in data_to_print:
            text = ''
            for i in row:
                text += i + ' ' * (10 - len(i))
            print(text)

    def _intersect(self, lst):
        if not lst:
            return [1] * self.signs_count
        res = []
        for i in range(len(lst[0])):
            if sum([lst[j][i] for j in range(len(lst))]) == len(lst):
                res.append(1)
            else:
                res.append(0)
        return res

    def _union(self, lst):
        if not lst:
            return [0] * self.signs_count
        res = []
        for i in range(len(lst[0])):
            if sum([lst[j][i] for j in range(len(lst))]) > 0:
                res.append(1)
            else:
                res.append(0)
        return res

    def _list_from_chars(self, chars):
        res = []
        for i in range(self.signs_count):
            if self.chars[i] in chars:
                res.append(1)
            else:
                res.append(0)
        return res

    def _chars_from_str(self, string):
        res = []
        for i in self.chars:
            if i in string:
                res.append(i)
        return res

    def _chars_from_list(self, lst):
        res = ""
        for i in range(len(lst)):
            if lst[i]:
                res += self.chars[i]
        return res if res else 'ø'

    def _get_objects_by_signs(self, signs_indexes):
        objects = []
        for j in range(len(self.data)):
            if sum([self.data[j][i] for i in signs_indexes]) == len(signs_indexes):
                objects.append(j)
        return objects

    def _get_objects_by_char(self, char):
        index = self.chars.index(char)
        return self._get_objects_by_signs([index])

    def _get_objects_by_x(self, x):
        signs_indexes = []
        for i in range(len(self.chars)):
            if x.find(self.chars[i]) != -1:
                signs_indexes.append(i)
        return self._get_objects_by_signs(signs_indexes)

    def _get_x_2_lst_from_objects(self, objects):
        obj_lists = [self.data[i] for i in range(self.objects_count) if i in objects]
        x_2_lst = self._intersect(obj_lists)
        return x_2_lst

    @property
    def chars(self):
        chars = []
        for i in range(self.signs_count):
            char = chr(ord('A') + i // 2)
            if i % 2 == 1:
                char = f'`{char.lower()}'
            chars.append(char)
        return chars

    def _make_prom_tab(self):
        prom_tab = []
        for char in self.chars:
            lst = self._list_from_chars(char)
            x_1 = self._get_objects_by_char(char)
            x_2_lst = self._get_x_2_lst_from_objects(x_1)
            x_2 = self._chars_from_list(x_2_lst)
            prom_tab.append(PromRow(char, lst, x_2, x_1, x_2, x_2_lst,))
        return prom_tab

    def _increment_iterator(self):
        bin_it = int(self.it, 2)
        bin_it += 1
        it = str(bin(bin_it))[2:]
        self.it = '0' * (self.signs_count - len(it)) + it

    def _approx(self, x: str):
        signs_x_2 = [prom_row.X_2_lst for prom_row in self.prom_tab if x.find(prom_row.X_name) != -1]
        return self._chars_from_list(self._union(signs_x_2))

    def _pass_no_keys_iterator(self):
        one_indexes = [i for i in range(len(self.it)) if self.it[i] == '1']
        last_one_index = max(one_indexes) if one_indexes else len(self.it)-1
        while self.it[last_one_index] == '1':
            self._increment_iterator()

    @staticmethod
    def _get_sublist(lst):
        return [[j for j in lst if j != i] for i in lst]

    def _is_equal_names(self, name1, name2):
        return set(self._chars_from_str(name1)) == set(self._chars_from_str(name2))

    def _name1_contain_name2(self, name1, name2):
        for char in self._chars_from_str(name2):
            if char not in name1:
                return False
        return True

    def _get_row_by_name(self, row_name):
        all_known_rows = self.prom_tab + self.result
        for row in all_known_rows:
            if self._is_equal_names(row.X_name, row_name):
                return row

    def _is_key(self, x: str):
        x_chars = self._chars_from_str(x)
        if len(x_chars) == 1:
            return True
        chars_subsets = self._get_sublist(x_chars)
        for chars_subset in chars_subsets:
            substring = ''.join(chars_subset)
            if not self._is_key(substring):
                return False
            substring_row = self._get_row_by_name(substring)
            if self._name1_contain_name2(substring_row.X_Apr, x):
                return False
            if self._name1_contain_name2(substring_row.X_2, x):
                return False
        return True

    def _conflict_trigger(self, x):
        for i in range(self.signs_count):
            if i % 2 == 0 and x.find(self.chars[i]) != -1 and x.find(self.chars[i+1]) != -1:
                return True
        return False

    def gen_next(self):
        bin_it = [int(i) for i in self.it]
        x = self._chars_from_list(bin_it)
        if self._is_key(x):
            x_1 = self._get_objects_by_x(x)
            x_2_lst = self._get_x_2_lst_from_objects(x_1)
            x_2 = self._chars_from_list(x_2_lst)
            self.result.append(
                PromRow(
                    x,
                    bin_it,
                    self._approx(x),
                    x_1,
                    x_2,
                    x_2_lst,
                )
            )
        if not self._is_key(x) or self._conflict_trigger(x):
            self._pass_no_keys_iterator()
        else:
            self._increment_iterator()

    def gen_all(self):
        self.it = '0' * self.signs_count
        while True:
            if len(self.it) > self.signs_count:
                break
            self.gen_next()

    def save_all_to_excel(self):
        wb = openpyxl.Workbook()
        ws1 = wb.get_active_sheet()
        ws1.title = 'Input'
        ws1.append(['G\\M'] + self.chars)
        for i in range(self.objects_count):
            ws1.append([i] + self.data[i])
        ws2 = wb.create_sheet('Auxiliary')
        ws2.append(['X', 'K', 'X+', 'X\'', 'X"'])
        for i in self.prom_tab:
            ws2.append([i.X_name, '+' if i.Key else '-', i.X_Apr, ''.join(map(str, i.X_1)) if i.X_1 else 'ø',
                        i.X_2 if i.X_2 else 'ø'])
        ws3 = wb.create_sheet('Output')
        ws3.append(['X', 'X', 'K', 'X+', 'X\'', 'X"'])
        for i in self.result:
            ws3.append([''.join(map(str, i.X_lst)), i.X_name, '+' if i.Key else '-',
                        i.X_Apr, ''.join(map(str, i.X_1)) if i.X_1 else 'ø', i.X_2 if i.X_2 else 'ø'])
        wb.save('output.xlsx')
